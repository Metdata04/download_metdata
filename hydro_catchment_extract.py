import os
import pdfplumber
import pandas as pd
from datetime import datetime

# Define the list of hydro catchment stations (fixed)
stations = [
    'Castlereigh', 'Norton', 'Maussakele', 'Canyon', 'Lakshapana',
    'Upper Kotmale', 'Victoriya', 'Kotmale', 'Randenigala', 'Rantambe',
    'Bowatenna', 'Ukuwela', 'Samanala Wawa', 'Kukuleganaga', 'Maskeliya', 'Neboda'
]

# Function to extract hydro catchment data from the PDF
def extract_hydro_catchment_data(pdf_path, pdf_missing=False):
    if pdf_missing:
        # If PDF is missing, fill all stations with 'NA' for that day
        rainfall_data = {station: 'NA' for station in stations}
    else:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                page = pdf.pages[0]  # Assuming the data is on the first page
                text = page.extract_text()

                # Split the text by new lines
                lines = text.split('\n')

                # Dictionary to hold rainfall data with station names as keys
                rainfall_data = {station: 'NA' for station in stations}

                # Search and extract the rainfall data based on station names
                for line in lines:
                    for station in stations:
                        if station in line:
                            # Extract the part of the line after the station name
                            parts = line.split(station)
                            if len(parts) > 1:
                                value_part = parts[1].strip()
                                # Extract the rainfall value (first word after the station name)
                                value = value_part.split()[0] if value_part else 'NA'
                                # Try to convert to float, otherwise leave as 'NA'
                                try:
                                    rainfall_data[station] = float(value)
                                except ValueError:
                                    rainfall_data[station] = value  # Keep it as a string if not convertible
                                 
        except Exception as e:
            print(f"Error reading PDF: {e}")
            rainfall_data = {station: 'NA' for station in stations}  # If PDF error, use 'NA'

    # Create a DataFrame with the date and the extracted rainfall data
    current_date = datetime.now().strftime('%m/%d/%Y')
    df = pd.DataFrame([rainfall_data])
    df.insert(0, 'Date', current_date)  # Add the Date column at the start

    return df


if __name__ == "__main__":
    date_string = datetime.now().strftime('%Y-%m-%d')
    pdf_path = f'metdata/daily_climate_update_{date_string}.pdf'  

    # Check if the PDF exists, otherwise set pdf_missing to True
    pdf_missing = not os.path.exists(pdf_path)

    hydro_data_df = extract_hydro_catchment_data(pdf_path, pdf_missing)

    # Save the extracted data to an Excel file
    os.makedirs('extracted_data', exist_ok=True)
    excel_file_path = os.path.join('extracted_data', 'hydro_catchment_data.xlsx')

    # Append data to the existing sheet or create a new file
    sheet_name = 'Data'  # Specify the sheet name
    if os.path.exists(excel_file_path):
        from openpyxl import load_workbook

        # Load the workbook to determine the last row in the sheet
        wb = load_workbook(excel_file_path)
        if sheet_name in wb.sheetnames:
            startrow = wb[sheet_name].max_row
        else:
            startrow = 0
        wb.close()

        # Append data to the sheet
        with pd.ExcelWriter(excel_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            hydro_data_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    else:
        # Create a new Excel file with the specified sheet
        hydro_data_df.to_excel(excel_file_path, sheet_name=sheet_name, index=False)

    print(f"Hydro catchment data extracted and saved to '{excel_file_path}'.")
